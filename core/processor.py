import os
import sys
import logging
import pandas as pd
from datetime import datetime
import tempfile
from pathlib import Path
import json
import time
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
import shutil
from PIL import Image as PILImage
import re
import io

# Add parent directory to path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

# Import utils modules directly
from utils import config_manager
from utils import excel_utils
from utils import image_utils

# Import get_downloads_folder from config_manager
from utils.config_manager import get_downloads_folder

# Setup logging
logger = logging.getLogger(__name__)
# <<< Тест 1: Проверяем, работает ли логгер этого модуля >>>
logger.critical("--- Logger for core.processor initialized ---") 

# <<< Constants for image fitting >>>
DEFAULT_CELL_WIDTH_PX = 300  # Ширина ячейки по умолчанию в пикселях
DEFAULT_CELL_HEIGHT_PX = 120  # Высота ячейки по умолчанию в пикселях
DEFAULT_IMG_QUALITY = 90
MIN_IMG_QUALITY = 1  # Снижено с 5% до 1% для еще большего сжатия
MIN_KB_PER_IMAGE = 10
MAX_KB_PER_IMAGE = 2048 # 2MB max per image, prevents extreme cases
SIZE_BUDGET_FACTOR = 0.85 # Use 85% of total size budget for images
ROW_HEIGHT_PADDING = 1 # Минимальный отступ для высоты строки
MIN_ASPECT_RATIO = 0.5 # Минимальное соотношение сторон (высота/ширина)
MAX_ASPECT_RATIO = 2.0 # Максимальное соотношение сторон (высота/ширина)
EXCEL_WIDTH_TO_PIXEL_RATIO = 7.0  # Коэффициент преобразования единиц Excel в пиксели
EXCEL_PX_TO_PT_RATIO = 0.75  # Коэффициент преобразования пикселей в единицы Excel
DEFAULT_EXCEL_COLUMN_WIDTH = 40  # Ширина колонки в единицах Excel (примерно 300px)
MIN_COLUMN_WIDTH_PX = 100  # Минимальная допустимая ширина колонки в пикселях

def ensure_temp_dir(prefix: str = "") -> str:
    """
    Создает и возвращает путь к временной директории.
    
    Args:
        prefix (str): Префикс для имени временной директории
    
    Returns:
        Путь к временной директории
    """
    temp_dir = os.path.join(tempfile.gettempdir(), f"{prefix}excelwithimages")
    os.makedirs(temp_dir, exist_ok=True)
    return temp_dir

def process_excel_file(
    file_path: str,
    article_col_name: str,
    image_folder: str,
    image_col_name: str = None,
    output_folder: str = None,
    max_total_file_size_mb: int = 100,
    progress_callback: callable = None,
    config: dict = None,
    header_row: int = 0,
    sheet_name: str = None,  # Добавляем параметр для имени листа
    secondary_image_folder: str = None,  # Папка с запасными изображениями (второй приоритет)
    tertiary_image_folder: str = None,   # Папка с дополнительными запасными изображениями (третий приоритет)
    output_filename: str = None  # Имя выходного файла
) -> Tuple[str, Optional[pd.DataFrame], int, Dict[str, List[str]], List[str], List[Dict]]:
    """
    Обрабатывает Excel файл, вставляя изображения на основе номеров артикулов.
    
    Args:
        file_path (str): Путь к Excel файлу
        article_col_name (str): Имя столбца с артикулами (или буква столбца)
        image_folder (str): Путь к папке с изображениями
        image_col_name (str, optional): Имя столбца для вставки изображений (или буква). По умолчанию None
        output_folder (str, optional): Папка для сохранения результата. По умолчанию None
        max_total_file_size_mb (int, optional): Макс. размер файла в МБ. По умолчанию 100 МБ
        progress_callback (callable, optional): Функция для отображения прогресса. По умолчанию None
        config (dict, optional): Словарь с настройками. По умолчанию None
        header_row (int, optional): Номер строки заголовка (0-based). По умолчанию 0
        sheet_name (str, optional): Имя листа Excel для обработки. По умолчанию None (первый лист)
        secondary_image_folder (str, optional): Путь к папке с запасными изображениями. По умолчанию None
        tertiary_image_folder (str, optional): Путь к дополнительной папке с запасными изображениями. По умолчанию None
        output_filename (str, optional): Имя выходного файла. По умолчанию None
    
    Returns:
        Tuple[str, pd.DataFrame, int, Dict[str, List[str]], List[str], List[Dict]]: 
            - Путь к файлу результата
            - DataFrame с данными
            - Количество вставленных изображений
            - Словарь с артикулами, для которых найдено несколько изображений (ключ: артикул, значение: список путей)
            - Список артикулов, для которых не найдены изображения
            - Список результатов поиска изображений (словари с информацией о поиске)
    """
    # <<< Используем print в stderr вместо logger >>>
    print(">>> ENTERING process_excel_file <<<\n", file=sys.stderr)
    sys.stderr.flush()
    
    print(f"[PROCESSOR] Начало обработки: {file_path}", file=sys.stderr)
    print(f"[PROCESSOR] Параметры: article_col={article_col_name}, img_folder={image_folder}, img_col={image_col_name}, max_total_mb={max_total_file_size_mb}, sheet_name={sheet_name}", file=sys.stderr)

    # --- Валидация входных данных ---
    # Проверка валидности обозначений колонок
    try:
        # Принимаем только буквенные обозначения колонок (A, B, C...)
        if not (article_col_name.isalpha() and image_col_name.isalpha()):
            err_msg = f"Неверное обозначение колонки: '{article_col_name}' или '{image_col_name}'. Используйте только буквенные обозначения (A, B, C...)"
            print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
            raise ValueError(err_msg)
            
        article_col_idx = excel_utils.column_letter_to_index(article_col_name)
        image_col_idx = excel_utils.column_letter_to_index(image_col_name)
    except Exception as e:
        err_msg = f"Неверное обозначение колонки: '{article_col_name}' или '{image_col_name}'. Ошибка: {str(e)}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise ValueError(err_msg)
        
    if not os.path.exists(file_path):
        err_msg = f"Файл не найден: {file_path}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise FileNotFoundError(err_msg)
    
    if not os.path.exists(image_folder):
        err_msg = f"Папка с изображениями не найдена: {image_folder}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise FileNotFoundError(err_msg)

    # --- Чтение Excel ---
    try:
        # Если указан конкретный лист, читаем его
        if sheet_name:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                engine='openpyxl',
                skiprows=0,
                header=None
            )
            print(f"[PROCESSOR] Excel-файл прочитан в DataFrame (sheet={sheet_name}, header=None). Строк данных: {len(df)}", file=sys.stderr)
        else:
            df = pd.read_excel(file_path, header=0, engine='openpyxl') 
            print(f"[PROCESSOR] Excel-файл прочитан в DataFrame (header=0). Строк данных: {len(df)}", file=sys.stderr)
        
        # --- Загрузка книги openpyxl ---
        wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=False)
        try:
            # Проверяем наличие листов в книге
            if not wb.sheetnames:
                print("[PROCESSOR ERROR] В файле нет листов для обработки.", file=sys.stderr)
                raise ValueError("Excel-файл не содержит листов. Пожалуйста, выберите файл с данными.")
                
            # Фильтруем листы, исключая листы с макросами
            valid_sheets = [sheet_name for sheet_name in wb.sheetnames if not sheet_name.startswith('xl/macrosheets/')]
            if not valid_sheets:
                print("[PROCESSOR ERROR] В файле нет обычных листов, только макросы.", file=sys.stderr)
                raise ValueError("Внимание! Этот файл Excel содержит только макросы, а не обычные таблицы данных. Пожалуйста, выберите файл Excel с обычными листами, содержащими таблицы с артикулами и данными для обработки.")
            
            # Если указан лист, выбираем его, иначе используем активный
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    print(f"[PROCESSOR] Работаем с указанным листом: {sheet_name}", file=sys.stderr)
                else:
                    print(f"[PROCESSOR ERROR] Указанный лист {sheet_name} не найден в файле. Доступные листы: {wb.sheetnames}", file=sys.stderr)
                    raise ValueError(f"Лист '{sheet_name}' не найден в файле. Доступные листы: {wb.sheetnames}")
            else:
                # Используем первый лист
                ws = wb.active
                print(f"[PROCESSOR] Загружена рабочая книга, работаем с активным листом: {ws.title}", file=sys.stderr)
        except Exception as e:
            print(f"[PROCESSOR ERROR] Ошибка при выборе листа: {e}", file=sys.stderr)
            # Делаем сообщение об ошибке более понятным для пользователя
            if "'dict' object has no attribute 'shape'" in str(e):
                raise ValueError("Выбранный лист не содержит табличных данных. Пожалуйста, выберите лист с необходимыми данными.")
            else:
                raise ValueError(f"Ошибка при выборе листа: {e}")
        
    except Exception as e:
        err_msg = f"Ошибка при чтении Excel-файла: {e}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        # Выводим traceback в консоль
        import traceback
        traceback.print_exc(file=sys.stderr)
        
        # Делаем сообщение об ошибке более понятным для пользователя
        user_friendly_msg = err_msg
        if "'dict' object has no attribute 'shape'" in str(e):
            user_friendly_msg = "Выбранный лист не содержит табличных данных. Пожалуйста, выберите лист с необходимыми данными."
        elif "No sheet" in str(e) or "not found" in str(e):
            user_friendly_msg = "Указанный лист не найден в файле. Пожалуйста, выберите существующий лист."
        elif "Empty" in str(e) or "no data" in str(e):
            user_friendly_msg = "Выбранный лист не содержит данных. Пожалуйста, выберите лист с данными."
            
        raise RuntimeError(user_friendly_msg) from e

    if df.empty:
        err_msg = "Excel-файл не содержит данных"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise ValueError(err_msg)

    # --- Проверка существования колонки артикулов ---
    # Преобразуем букву колонки в индекс
    article_col_idx = excel_utils.column_letter_to_index(article_col_name)
    article_col_name = df.columns[article_col_idx]
    
    # Принудительно конвертируем значения артикулов в строковый тип
    df[article_col_name] = df[article_col_name].astype(str)
    
    articles = df[article_col_name].tolist()
    print(f"[PROCESSOR] Получено {len(articles)} артикулов из колонки {article_col_name}", file=sys.stderr)
    
    if article_col_name not in df.columns:
        err_msg = f"Колонка с артикулами '{article_col_name}' не найдена в файле. Доступные колонки: {list(df.columns)}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise ValueError(err_msg)
    print(f"[PROCESSOR] Колонка с артикулами: '{article_col_name}'", file=sys.stderr)

    # --- Определение КОЛИЧЕСТВА строк с НЕНУЛЕВЫМИ артикулами для расчета лимита ---
    # Считаем строки, где артикул не пустой
    non_empty_article_rows = df[article_col_name].notna() & (df[article_col_name].astype(str).str.strip() != '')
    article_count = non_empty_article_rows.sum()
    
    if article_count == 0:
        article_count = 1 # Избегаем деления на ноль
        print("[PROCESSOR WARNING] Не найдено строк с непустыми артикулами для расчета лимита размера изображения. Используется значение по умолчанию.", file=sys.stderr)
    else:
        print(f"[PROCESSOR] Найдено {article_count} строк с непустыми артикулами.", file=sys.stderr)
        
    # --- Расчет лимита размера на одно изображение ---
    image_size_budget_mb = max_total_file_size_mb * SIZE_BUDGET_FACTOR
    target_kb_per_image = (image_size_budget_mb * 1024) / article_count if article_count > 0 else MAX_KB_PER_IMAGE
    target_kb_per_image = max(MIN_KB_PER_IMAGE, min(target_kb_per_image, MAX_KB_PER_IMAGE)) 
    print(f"[PROCESSOR] Расчетный лимит размера на изображение: {target_kb_per_image:.1f} КБ", file=sys.stderr)

    # --- Подготовка папки для обработанных изображений ---
    temp_image_dir_created = False
    if not image_folder:
        image_folder = ensure_temp_dir("processed_images_")
        temp_image_dir_created = True
        print(f"[PROCESSOR] Создана временная директория для обработанных изображений: {image_folder}", file=sys.stderr)
    elif not os.path.exists(image_folder):
         os.makedirs(image_folder)
         print(f"[PROCESSOR] Создана папка для обработанных изображений: {image_folder}", file=sys.stderr)


    # --- Подготовка к вставке изображений ---
    try:
        # НАПРЯМУЮ ИСПОЛЬЗУЕМ УКАЗАННУЮ БУКВУ КОЛОНКИ
        image_col_letter_excel = image_col_name
        print(f"[PROCESSOR] Изображения будут вставляться в колонку: '{image_col_letter_excel}'", file=sys.stderr)
    except Exception as e:
         err_msg = f"Ошибка при подготовке колонки для изображений ('{article_col_name}'): {e}"
         print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
         import traceback
         traceback.print_exc(file=sys.stderr)
         raise RuntimeError(err_msg) from e

    # --- Настройка ШИРИНЫ КОЛОНКИ ---
    try:
        # ВАЖНО: НЕ устанавливаем ширину колонки вручную, используем текущую ширину Excel
        # Проверяем фактическую ширину колонки
        column_width_excel = ws.column_dimensions[image_col_letter_excel].width
        
        # Если ширина колонки не определена, используем стандартную ширину Excel (8.43)
        if not column_width_excel:
            column_width_excel = ws.sheet_format.defaultColWidth or 8.43
        
        # Переводим в пиксели для информации
        actual_width_px = int(column_width_excel * EXCEL_WIDTH_TO_PIXEL_RATIO)
        print(f"[PROCESSOR] Фактическая ширина столбца {image_col_letter_excel}: {column_width_excel:.2f} ед. Excel (≈ {actual_width_px} пикс.)", file=sys.stderr)
    except Exception as e:
        print(f"[PROCESSOR WARNING] Не удалось определить ширину столбца {image_col_letter_excel}: {e}", file=sys.stderr)


    # --- Обработка строк и вставка изображений ---
    images_inserted = 0
    rows_processed = 0
    total_processed_image_size_kb = 0
    
    # Создаем список для хранения результатов поиска изображений
    image_search_results = []
    
    # Инициализируем списки для хранения результатов
    not_found_articles = []
    multiple_images_found = {}
    
    print("[PROCESSOR] --- Начало итерации по строкам DataFrame ---", file=sys.stderr)
    
    # Переменные для определения оптимального качества сжатия
    first_image_processed = False
    successful_quality = DEFAULT_IMG_QUALITY  # Если не найдено, используем значение по умолчанию
    quality_determined = False  # Флаг, указывающий, был ли определен уровень качества
    
    # Итерация по строкам таблицы
    for excel_row_index, row in df.iterrows():
        # Проверяем, нужно ли обновить прогресс
        if progress_callback and excel_row_index % 5 == 0:  # Обновление каждые 5 строк
            progress_value = min(0.9, (excel_row_index / len(df)) * 0.9)  # 90% прогресса на обработку строк
            progress_callback(progress_value, f"Обработка строки {excel_row_index + 1} из {len(df)}")
        
        rows_processed += 1
        
        # Сначала получаем артикул
        article_str = str(row[article_col_name]).strip()
        
        print(f"[PROCESSOR] Обработка строки {excel_row_index}, артикул: '{article_str}'", file=sys.stderr)
        
        if pd.isna(row[article_col_name]) or article_str.strip() == "":
            print(f"[PROCESSOR]   Пустой артикул в строке {excel_row_index}, пропускаем", file=sys.stderr)
            continue
        
        # Find images for this article in multiple folders
        supported_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')
        
        # Получаем пути к резервным папкам из параметров или конфигурации
        secondary_folder_path = secondary_image_folder or config_manager.get_setting("paths.secondary_images_folder_path", "")
        tertiary_folder_path = tertiary_image_folder or config_manager.get_setting("paths.tertiary_images_folder_path", "")
        
        # Логируем папки для диагностики
        print(f"[PROCESSOR DEBUG] Поиск изображений для артикула '{article_str}' в папках:", file=sys.stderr)
        print(f"[PROCESSOR DEBUG]   Основная: {image_folder}", file=sys.stderr)
        print(f"[PROCESSOR DEBUG]   Вторичная: {secondary_folder_path}", file=sys.stderr)
        print(f"[PROCESSOR DEBUG]   Третичная: {tertiary_folder_path}", file=sys.stderr)
        
        search_result = image_utils.find_images_in_multiple_folders(
            article_str, 
            image_folder,
            secondary_folder_path,
            tertiary_folder_path,
            supported_extensions,
            search_recursively=True
        )
        
        # Добавляем результат поиска в список
        search_result['row_index'] = excel_row_index
        search_result['article'] = article_str
        search_result['image_folders'] = {
            'primary': image_folder,
            'secondary': secondary_folder_path,
            'tertiary': tertiary_folder_path
        }
        image_search_results.append(search_result)
        
        # If no images found, record and continue
        if not search_result["found"]:
            print(f"[PROCESSOR WARNING]   Для артикула '{article_str}' (строка {excel_row_index}) не найдено изображений. Пропускаем.", file=sys.stderr)
            # Добавляем артикул в список не найденных
            not_found_articles.append(article_str)
            continue
        
        # If multiple images found, record for report
        all_image_paths = search_result["images"]
        source_folder_priority = search_result["source_folder"]
        
        if len(all_image_paths) > 1:
            print(f"[PROCESSOR INFO]   Найдено несколько изображений для артикула '{article_str}': {len(all_image_paths)}", file=sys.stderr)
            multiple_images_found[article_str] = all_image_paths
            # Still proceed with the first image
        
        image_path = all_image_paths[0]
        print(f"[PROCESSOR]   Выбрано первое найденное изображение: {image_path} (папка приоритета {source_folder_priority})", file=sys.stderr)

        # Проверяем, удовлетворяет ли изображение требованиям по размеру
        original_size_kb = os.path.getsize(image_path) / 1024
        print(f"[PROCESSOR]   Размер исходного изображения: {original_size_kb:.1f} КБ, лимит: {target_kb_per_image:.1f} КБ", file=sys.stderr)
        
        # 1. ОПТИМИЗАЦИЯ ИЗОБРАЖЕНИЯ (если требуется)
        optimized_buffer = None
        
        if original_size_kb <= target_kb_per_image:
            # Если размер уже подходит, просто загружаем изображение без оптимизации
            print(f"[PROCESSOR]   Изображение уже удовлетворяет требованиям по размеру, загружаем без оптимизации", file=sys.stderr)
            try:
                with open(image_path, 'rb') as f_orig:
                    optimized_buffer = io.BytesIO(f_orig.read())
                print(f"[PROCESSOR]   Загружено без оптимизации, размер: {optimized_buffer.tell()/1024:.1f} КБ", file=sys.stderr)
                optimized_buffer.seek(0)
            except Exception as e:
                print(f"[PROCESSOR ERROR]   Ошибка при загрузке изображения без оптимизации: {e}", file=sys.stderr)
                # Если не удалось загрузить, попробуем оптимизировать
        else:
            # Требуется оптимизация
            print(f"[PROCESSOR]   Вызов optimize_image_for_excel для {image_path} с лимитом {target_kb_per_image:.1f} КБ", file=sys.stderr)
            
            try:
                # Если уровень качества еще не определен, определяем его на первом изображении,
                # которое требует оптимизации
                if not quality_determined:
                    # Ищем оптимальное качество для сжатия
                    print(f"[PROCESSOR]   ОПРЕДЕЛЕНИЕ ОПТИМАЛЬНОГО КАЧЕСТВА: поиск качества от {DEFAULT_IMG_QUALITY}% до {MIN_IMG_QUALITY}%", file=sys.stderr)
                    optimized_buffer = image_utils.optimize_image_for_excel(
                        image_path, 
                        target_size_kb=target_kb_per_image,
                        quality=DEFAULT_IMG_QUALITY,
                        min_quality=MIN_IMG_QUALITY    
                    )
                    
                    # Помечаем что определили качество
                    quality_determined = True
                    
                    # После успешной оптимизации первого изображения определяем качество,
                    # которое лучше всего подошло
                    if optimized_buffer and optimized_buffer.getbuffer().nbytes > 0:
                        # Мы можем полагаться только на лог-сообщения для определения качества
                        # Находим в логах последнюю строку с "Итоговое качество сжатия"
                        try:
                            # Сначала попытаемся прочитать текущий stderr буфер, если это возможно
                            quality_found = False
                            
                            # Попытка найти в логах
                            with open("logs/app_latest.log", "r", encoding="utf-8") as log_file:
                                log_lines = log_file.readlines()
                                # Ищем два типа строк с информацией о качестве
                                quality_pattern1 = re.compile(r'\[optimize_excel\] Итоговое качество сжатия: (\d+)%')
                                quality_pattern2 = re.compile(r'-> Успех! Размер .* c качеством (\d+)')
                                # Добавляем поиск специального маркера
                                quality_marker = re.compile(r'\[QUALITY_MARKER\] НАЙДЕНО_КАЧЕСТВО_ДЛЯ_ИЗОБРАЖЕНИЯ: (\d+)')
                                
                                # Сначала ищем в последних 100 строках (самая свежая информация)
                                recent_lines = log_lines[-100:] if len(log_lines) > 100 else log_lines
                                
                                for line in reversed(recent_lines):
                                    # Приоритетно ищем специальный маркер
                                    match = quality_marker.search(line)
                                    if match:
                                        found_quality = int(match.group(1))
                                        successful_quality = found_quality
                                        print(f"[PROCESSOR]   НАЙДЕН СПЕЦИАЛЬНЫЙ МАРКЕР КАЧЕСТВА: {successful_quality}%", file=sys.stderr)
                                        quality_found = True
                                        break
                                        
                                    match = quality_pattern1.search(line)
                                    if match:
                                        found_quality = int(match.group(1))
                                        successful_quality = found_quality
                                        print(f"[PROCESSOR]   Определено оптимальное качество из последних логов (pattern 1): {successful_quality}%", file=sys.stderr)
                                        quality_found = True
                                        break
                                        
                                    match = quality_pattern2.search(line)
                                    if match:
                                        found_quality = int(match.group(1))
                                        successful_quality = found_quality
                                        print(f"[PROCESSOR]   Определено оптимальное качество из последних логов (pattern 2): {successful_quality}%", file=sys.stderr)
                                        quality_found = True
                                        break
                                
                                # Если не нашли в последних строках, ищем во всем файле
                                if not quality_found:
                                    print(f"[PROCESSOR]   Качество не найдено в последних строках логов, ищем во всем файле...", file=sys.stderr)
                                    for line in reversed(log_lines):
                                        match = quality_pattern1.search(line)
                                        if match:
                                            found_quality = int(match.group(1))
                                            successful_quality = found_quality
                                            print(f"[PROCESSOR]   Определено оптимальное качество из всех логов: {successful_quality}%", file=sys.stderr)
                                            quality_found = True
                                            break
                                
                                # Если все еще не нашли, попробуем прочитать из временного файла
                                if not quality_found:
                                    try:
                                        temp_quality_file = os.path.join(tempfile.gettempdir(), "last_image_quality.txt")
                                        if os.path.exists(temp_quality_file):
                                            with open(temp_quality_file, "r") as qf:
                                                file_quality = int(qf.read().strip())
                                                successful_quality = file_quality
                                                print(f"[PROCESSOR]   Определено качество из временного файла: {successful_quality}%", file=sys.stderr)
                                                quality_found = True
                                    except Exception as file_err:
                                        print(f"[PROCESSOR]   Ошибка чтения временного файла с качеством: {file_err}", file=sys.stderr)
                                
                                # Если все еще не нашли, используем мин. значение
                                if not quality_found:
                                    successful_quality = MIN_IMG_QUALITY  # Прямое использование мин. качества
                                    print(f"[PROCESSOR]   Качество не найдено в логах. Используем минимальное значение: {successful_quality}%", file=sys.stderr)
                        except Exception as log_e:
                            print(f"[PROCESSOR WARNING]   Ошибка при чтении лог-файла: {log_e}. Используем минимальное качество.", file=sys.stderr)
                            successful_quality = MIN_IMG_QUALITY  # Минимальное качество без добавления
                    
                    # Сообщаем о выбранном качестве для всех последующих изображений
                    print(f"[PROCESSOR]   ВАЖНО: Для всех последующих изображений будет использовано качество {successful_quality}%", file=sys.stderr)
                else:
                    # Для всех последующих изображений используем найденное качество
                    print(f"[PROCESSOR]   Используем найденное качество {successful_quality}% для изображения {image_path}", file=sys.stderr)
                    optimized_buffer = image_utils.optimize_image_for_excel(
                        image_path, 
                        target_size_kb=target_kb_per_image,
                        quality=successful_quality,  # Начальное = найденное качество 
                        min_quality=successful_quality  # Мин. качество = найденное качество (без итераций)
                    )
            except Exception as e:
                print(f"[PROCESSOR ERROR]   Ошибка при оптимизации изображения: {e}", file=sys.stderr)
                # Если не удалось оптимизировать, попробуем загрузить оригинальное изображение
                try:
                    with open(image_path, 'rb') as f_orig:
                        optimized_buffer = io.BytesIO(f_orig.read())
                    print(f"[PROCESSOR]   Загружен оригинал из-за ошибки оптимизации, размер: {optimized_buffer.tell()/1024:.1f} КБ", file=sys.stderr)
                    optimized_buffer.seek(0)
                except Exception as load_e:
                    print(f"[PROCESSOR ERROR]   Не удалось загрузить оригинальное изображение: {load_e}", file=sys.stderr)
                    continue
        
        if optimized_buffer and optimized_buffer.getbuffer().nbytes > 0:
            buffer_size_kb = optimized_buffer.tell() / 1024
            print(f"[PROCESSOR]   Размер буфера для вставки: {buffer_size_kb:.1f} КБ", file=sys.stderr)
            current_image_size_kb = buffer_size_kb
            total_processed_image_size_kb += current_image_size_kb
            
            # Дополнительная проверка буфера - убеждаемся, что это действительно изображение
            optimized_buffer.seek(0)
            try:
                verification_img = PILImage.open(optimized_buffer)
                img_format = verification_img.format
                img_width_px, img_height_px = verification_img.size
                print(f"[PROCESSOR]   ВЕРИФИКАЦИЯ: буфер содержит изображение формата {img_format}, {img_width_px}x{img_height_px}", file=sys.stderr)
                
                # Создаем временную копию буфера для сохранения в файл (для отладки)
                try:
                    debug_copy = io.BytesIO(optimized_buffer.getvalue())
                    temp_debug_path = os.path.join(tempfile.gettempdir(), f"debug_image_{time.time()}.jpg")
                    with open(temp_debug_path, "wb") as debug_file:
                        debug_file.write(debug_copy.getvalue())
                    print(f"[PROCESSOR]   Создана отладочная копия изображения: {temp_debug_path}", file=sys.stderr)
                except Exception as debug_e:
                    print(f"[PROCESSOR]   Примечание: не удалось создать отладочную копию: {debug_e}", file=sys.stderr)
                
                # Сбрасываем указатель в начало буфера после верификации
                optimized_buffer.seek(0)
            except Exception as verify_e:
                print(f"[PROCESSOR] ОШИБКА ВЕРИФИКАЦИИ: Буфер не содержит корректного изображения: {verify_e}", file=sys.stderr)
                # Пробуем сохранить проблемный буфер для анализа
                try:
                    error_path = os.path.join(tempfile.gettempdir(), f"error_buffer_{time.time()}.bin")
                    with open(error_path, "wb") as error_file:
                        error_file.write(optimized_buffer.getvalue())
                    print(f"[PROCESSOR]   Сохранён проблемный буфер для анализа: {error_path}", file=sys.stderr)
                except Exception as err_save_e:
                    print(f"[PROCESSOR]   Не удалось сохранить проблемный буфер: {err_save_e}", file=sys.stderr)
                    
                # Если буфер некорректен, пробуем загрузить оригинальное изображение
                try:
                    print(f"[PROCESSOR]   Пробуем загрузить оригинальное изображение как резервный вариант", file=sys.stderr)
                    with open(image_path, "rb") as original_file:
                        optimized_buffer = io.BytesIO(original_file.read())
                    print(f"[PROCESSOR]   Загружено оригинальное изображение размером {optimized_buffer.getbuffer().nbytes / 1024:.1f} КБ", file=sys.stderr)
                    optimized_buffer.seek(0)
                    verification_img = PILImage.open(optimized_buffer)
                    img_width_px, img_height_px = verification_img.size
                except Exception as orig_load_e:
                    print(f"[PROCESSOR] КРИТИЧЕСКАЯ ОШИБКА: Не удалось загрузить даже оригинальное изображение: {orig_load_e}", file=sys.stderr)
                    continue  # Пропускаем эту итерацию
            
            # Получаем размеры изображения напрямую из буфера
            try:
                optimized_buffer.seek(0)
                img = PILImage.open(optimized_buffer)
                img_width_px, img_height_px = img.size
                print(f"[PROCESSOR]     Получены размеры из буфера: {img_width_px}x{img_height_px}", file=sys.stderr)
            except Exception as dim_e:
                print(f"[PROCESSOR] WARNING: Не удалось получить размеры изображения из буфера: {dim_e}", file=sys.stderr)
            
            # Вставляем изображение в Excel
            try:
                # Проверяем, что буфер изображения не пустой
                if not optimized_buffer or optimized_buffer.getbuffer().nbytes == 0:
                    print(f"[PROCESSOR WARNING] Пустой буфер изображения для артикула '{article_str}' (строка {excel_row_index})", file=sys.stderr)
                    continue
                
                # 1. Определяем фактическую ширину колонки Excel
                column_width_excel = None
                try:
                    # Получаем прямой доступ к размеру колонки
                    column_width_excel = ws.column_dimensions[image_col_letter_excel].width
                except Exception:
                    pass
                
                # Если ширина не определена, используем стандартную ширину листа
                if not column_width_excel:
                    column_width_excel = ws.sheet_format.defaultColWidth or 8.43  # Стандартный размер колонки Excel
                
                # Переводим в пиксели для расчета размеров изображения
                target_width_px = int(column_width_excel * EXCEL_WIDTH_TO_PIXEL_RATIO)
                print(f"[PROCESSOR] Используем фактическую ширину столбца {image_col_letter_excel}: {column_width_excel:.2f} ед. Excel ({target_width_px} пикс.)", file=sys.stderr)
                
                # Убираем корректировку - используем точную ширину столбца
                # 2. Получаем размеры исходного изображения для сохранения пропорций
                optimized_buffer.seek(0)
                pil_image = PILImage.open(optimized_buffer)
                img_width, img_height = pil_image.size
                aspect_ratio = img_height / img_width if img_width > 0 else 1.0
                optimized_buffer.seek(0)
                print(f"[PROCESSOR] Размеры оригинального изображения: {img_width}x{img_height}, соотношение сторон: {aspect_ratio:.2f}", file=sys.stderr)
                
                # Рассчитываем высоту изображения с сохранением пропорций
                target_height_px = int(target_width_px * aspect_ratio)
                
                # Формируем адрес ячейки для вставки
                anchor_cell = f"{image_col_letter_excel}{excel_row_index + 1 + header_row}"
                
                # Вставляем изображение с рассчитанными размерами
                print(f"[PROCESSOR] Вставляем изображение с размерами: {target_width_px}x{target_height_px} пикс.", file=sys.stderr)
                excel_utils.insert_image_from_buffer(
                    ws, 
                    optimized_buffer,
                    anchor_cell,
                    width=target_width_px,
                    height=target_height_px,
                    preserve_aspect_ratio=True
                )
                
                # 3. Устанавливаем высоту строки, чтобы изображение точно вписалось
                row_num = excel_row_index + 1 + header_row
                # Преобразуем пиксели в единицы Excel
                row_height_excel = target_height_px * EXCEL_PX_TO_PT_RATIO
                excel_utils.set_row_height(ws, row_num, row_height_excel)
                print(f"[PROCESSOR] Установлена высота строки {row_num}: {row_height_excel:.2f} ед. Excel для вмещения изображения", file=sys.stderr)
                
                # Увеличиваем счетчик успешно вставленных изображений
                images_inserted += 1
                print(f"[PROCESSOR] Изображение успешно вставлено в ячейку {anchor_cell}", file=sys.stderr)
                
            except Exception as e:
                print(f"[PROCESSOR ERROR] Ошибка при вставке изображения: {e}", file=sys.stderr)
                traceback.print_exc(file=sys.stderr)
                # Если количество вставленных изображений > 0, продолжаем
                if images_inserted > 0:
                    print(f"[PROCESSOR WARNING] Вставка изображения не удалась, но продолжаем обработку других строк", file=sys.stderr)
                    continue
                else:
                    # Это первое изображение и мы получили ошибку
                    print(f"[PROCESSOR ERROR] Критическая ошибка при вставке первого изображения: {e}", file=sys.stderr)
                    raise
        else:
            print(f"[PROCESSOR WARNING] Пустой буфер изображения для артикула '{article_str}' (строка {excel_row_index})", file=sys.stderr)
    
    # --- Сохранение результата ---
    print("\n[PROCESSOR] --- Сохранение результата ---", file=sys.stderr)
    
    try:
        # Создаем папку для результатов, если не существует
        if not output_folder:
            output_folder = os.path.join(os.path.dirname(file_path), "processed")
        
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
            print(f"[PROCESSOR] Создана папка для результатов: {output_folder}", file=sys.stderr)
        
        # Генерируем уникальное имя файла с датой и временем
        if output_filename:
            result_file_path = os.path.join(output_folder, output_filename)
        else:
            output_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_with Images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            result_file_path = os.path.join(output_folder, output_filename)
        
        # Сохраняем Excel-файл
        try:
            wb.save(result_file_path)
            print(f"[PROCESSOR] Результат сохранен в файл: {result_file_path}", file=sys.stderr)
            
            # Получаем фактический размер файла
            file_size_mb = os.path.getsize(result_file_path) / (1024 * 1024)
            print(f"[PROCESSOR] Фактический размер файла: {file_size_mb:.2f} МБ", file=sys.stderr)
            
            if progress_callback:
                progress_callback(1.0, f"Готово. Размер файла: {file_size_mb:.2f} MB")
        except Exception as save_e:
            print(f"[PROCESSOR] ОШИБКА ПРИ СОХРАНЕНИИ EXCEL: {save_e}", file=sys.stderr)
            # Вывод подробной ошибки в лог
            import traceback
            traceback.print_exc(file=sys.stderr)
            raise RuntimeError(f"Ошибка при сохранении файла: {save_e}")
    except Exception as out_e:
        print(f"[PROCESSOR] ОШИБКА ПРИ ПОДГОТОВКЕ ВЫВОДА: {out_e}", file=sys.stderr)
        raise RuntimeError(f"Ошибка при подготовке вывода: {out_e}")
    
    print(f"[PROCESSOR] СТАТИСТИКА: Обработано строк: {rows_processed}, вставлено изображений: {images_inserted}", file=sys.stderr)
    print(f"[PROCESSOR] Общий размер вставленных изображений: {total_processed_image_size_kb:.2f} КБ", file=sys.stderr)
    
    # Добавляем результаты поиска изображений к возвращаемым данным
    return result_file_path, df, images_inserted, multiple_images_found, not_found_articles, image_search_results

def get_column_width_pixels(ws, column_letter):
    """
    Получает фактическую ширину колонки в пикселях на основе настроек Excel.
    Не использует значений по умолчанию, берет только реальные данные из Excel.
    
    Args:
        ws: Рабочий лист Excel
        column_letter: Буква колонки (например, 'A', 'B', etc.)
        
    Returns:
        int: Ширина колонки в пикселях
    """
    try:
        # Получаем размер колонки из объекта column_dimensions
        column_dimensions = ws.column_dimensions.get(column_letter)
        
        # Проверяем, существует ли размер для данной колонки
        if column_dimensions and hasattr(column_dimensions, 'width') and column_dimensions.width is not None:
            width_in_excel_units = column_dimensions.width
            print(f"[PROCESSOR DEBUG] Получена ширина колонки {column_letter}: {width_in_excel_units} ед. Excel", file=sys.stderr)
        else:
            # Используем стандартную ширину из настроек листа
            width_in_excel_units = ws.sheet_format.defaultColWidth or 8.43  # Стандартный размер колонки Excel
            print(f"[PROCESSOR DEBUG] Используется стандартная ширина листа для колонки {column_letter}: {width_in_excel_units} ед. Excel", file=sys.stderr)
        
        # Преобразуем единицы Excel в пиксели
        pixels = int(width_in_excel_units * EXCEL_WIDTH_TO_PIXEL_RATIO)
        print(f"[PROCESSOR DEBUG] Ширина колонки {column_letter} в пикселях: {pixels} px", file=sys.stderr)
        return pixels
    except Exception as e:
        print(f"[PROCESSOR WARNING] Ошибка при получении ширины колонки {column_letter}: {e}", file=sys.stderr)
        # Используем стандартную ширину Excel в крайнем случае
        standard_width = 8.43  # Стандартная ширина колонки Excel
        return int(standard_width * EXCEL_WIDTH_TO_PIXEL_RATIO)