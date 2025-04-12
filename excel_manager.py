"""
Модуль для работы с Excel и добавления изображений в таблицы
"""
import os
import sys
import logging
import subprocess
import importlib.util
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple, Union

# Функция для проверки и установки необходимых библиотек
def check_and_install_dependencies():
    """
    Проверяет наличие необходимых библиотек и устанавливает их при необходимости
    """
    # Проверяем, был ли передан флаг, сигнализирующий о том, что мы уже пытались установить зависимости
    if "--deps_installed" in sys.argv or os.environ.get('EXCELWITHIMAGES_DEPS_INSTALLED') == '1':
        # Если флаг есть, просто выходим из функции, чтобы избежать зацикливания
        return True
    
    # Список необходимых библиотек
    required_packages = [
        "openpyxl",
        "Pillow"
    ]
    
    packages_to_install = []
    
    # Проверяем наличие каждой библиотеки
    for package in required_packages:
        try:
            # Пытаемся импортировать пакет напрямую, это более надежный способ
            if package == "Pillow":
                __import__("PIL")
            else:
                package_name = package.split(">=")[0]
                __import__(package_name)
        except ImportError:
            packages_to_install.append(package)
    
    # Если есть библиотеки для установки
    if packages_to_install:
        print("Обнаружены отсутствующие библиотеки. Начинаю установку...")
        
        for package in packages_to_install:
            print(f"Устанавливаю {package}...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", package])
                print(f"Библиотека {package} успешно установлена")
            except subprocess.CalledProcessError as e:
                print(f"Ошибка при установке библиотеки {package}: {e}")
                return False
        
        print("Все необходимые библиотеки установлены")
        
        # Устанавливаем переменную среды, чтобы избежать повторной установки
        os.environ['EXCELWITHIMAGES_DEPS_INSTALLED'] = '1'
        
        # Перезапускаем скрипт после установки библиотек с флагом, указывающим, что библиотеки уже были установлены
        print("Перезапуск скрипта...")
        new_args = sys.argv.copy()
        new_args.append("--deps_installed")
        os.execv(sys.executable, [sys.executable] + new_args)
    
    return True

# Проверяем и устанавливаем необходимые библиотеки до импорта модулей
if not check_and_install_dependencies():
    print("Возникли проблемы при установке необходимых библиотек. Программа будет завершена.")
    sys.exit(1)

# Удаляем флаг --deps_installed из аргументов, чтобы он не мешал обработке аргументов командной строки
if "--deps_installed" in sys.argv:
    sys.argv.remove("--deps_installed")

# Импортируем после проверки зависимостей
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

# Используем относительные импорты вместо абсолютных
try:
    from .utils.excel_utils import (
        open_workbook,
        save_workbook,
        get_cell_value,
        set_cell_value,
        find_column_by_header,
        set_column_width,
        apply_style_to_cell,
        create_table_from_data
    )
    from .utils.image_utils import (
        resize_image,
        create_excel_image,
        convert_pil_to_excel_image,
        find_images_in_directory,
        find_images_by_pattern,
        get_image_data
    )
except ImportError:
    # Если относительный импорт не работает (например, при прямом запуске файла),
    # пробуем абсолютный импорт
    from utils.excel_utils import (
        open_workbook,
        save_workbook,
        get_cell_value,
        set_cell_value,
        find_column_by_header,
        set_column_width,
        apply_style_to_cell,
        create_table_from_data
    )
    from utils.image_utils import (
        resize_image,
        create_excel_image,
        convert_pil_to_excel_image,
        find_images_in_directory,
        find_images_by_pattern,
        get_image_data
    )

logger = logging.getLogger(__name__)

class ExcelManager:
    """Класс для работы с Excel-файлами и добавления изображений"""
    
    def __init__(self, excel_file: str = None):
        """
        Инициализирует менеджер Excel
        
        Args:
            excel_file (str, optional): Путь к Excel-файлу. Если None, создается новый файл
        """
        self.excel_file = excel_file
        self.workbook = None
        self.active_sheet = None
        
        if excel_file and os.path.exists(excel_file):
            self.open_file(excel_file)
        else:
            self.create_new_workbook()
            
        logger.info(f"ExcelManager инициализирован {'с файлом ' + excel_file if excel_file else 'с новым файлом'}")
    
    def open_file(self, file_path: str) -> bool:
        """
        Открывает Excel-файл
        
        Args:
            file_path (str): Путь к файлу
            
        Returns:
            bool: True, если файл успешно открыт
        """
        try:
            self.workbook = open_workbook(file_path)
            if self.workbook:
                self.excel_file = file_path
                self.active_sheet = self.workbook.active
                logger.info(f"Excel-файл успешно открыт: {file_path}")
                return True
            return False
        except Exception as e:
            logger.error(f"Ошибка при открытии Excel-файла {file_path}: {e}")
            return False
    
    def create_new_workbook(self) -> bool:
        """
        Создает новую рабочую книгу Excel
        
        Returns:
            bool: True, если книга успешно создана
        """
        try:
            self.workbook = openpyxl.Workbook()
            self.active_sheet = self.workbook.active
            self.active_sheet.title = "Лист1"
            logger.info("Создана новая рабочая книга Excel")
            return True
        except Exception as e:
            logger.error(f"Ошибка при создании новой книги Excel: {e}")
            return False
    
    def save(self, file_path: str = None) -> bool:
        """
        Сохраняет рабочую книгу Excel
        
        Args:
            file_path (str, optional): Путь для сохранения. Если None, используется текущий путь
            
        Returns:
            bool: True, если книга успешно сохранена
        """
        if not self.workbook:
            logger.error("Нет активной книги для сохранения")
            return False
            
        save_path = file_path or self.excel_file
        if not save_path:
            logger.error("Не указан путь для сохранения Excel-файла")
            return False
            
        return save_workbook(self.workbook, save_path)
    
    def create_sheet(self, title: str) -> Optional[Worksheet]:
        """
        Создает новый лист в книге
        
        Args:
            title (str): Название листа
            
        Returns:
            Optional[Worksheet]: Объект листа или None в случае ошибки
        """
        try:
            if not self.workbook:
                logger.error("Нет активной книги для создания листа")
                return None
                
            sheet = self.workbook.create_sheet(title=title)
            logger.info(f"Создан новый лист: {title}")
            return sheet
        except Exception as e:
            logger.error(f"Ошибка при создании листа {title}: {e}")
            return None
    
    def set_active_sheet(self, sheet_name: str) -> bool:
        """
        Устанавливает активный лист по имени
        
        Args:
            sheet_name (str): Название листа
            
        Returns:
            bool: True, если лист успешно активирован
        """
        try:
            if not self.workbook:
                logger.error("Нет активной книги")
                return False
                
            if sheet_name in self.workbook.sheetnames:
                self.active_sheet = self.workbook[sheet_name]
                logger.info(f"Установлен активный лист: {sheet_name}")
                return True
            else:
                logger.error(f"Лист не найден: {sheet_name}")
                return False
        except Exception as e:
            logger.error(f"Ошибка при установке активного листа {sheet_name}: {e}")
            return False
    
    def add_image_to_cell(self, image_path: str, cell: str, 
                         max_width: int = 300, max_height: int = 200,
                         sheet: Worksheet = None) -> bool:
        """
        Добавляет изображение в указанную ячейку
        
        Args:
            image_path (str): Путь к изображению
            cell (str): Адрес ячейки (например, 'A1')
            max_width (int): Максимальная ширина изображения в пикселях
            max_height (int): Максимальная высота изображения в пикселях
            sheet (Worksheet, optional): Лист для добавления. Если None, используется активный лист
            
        Returns:
            bool: True, если изображение успешно добавлено
        """
        try:
            if not self.workbook:
                logger.error("Нет активной книги Excel")
                return False
                
            target_sheet = sheet or self.active_sheet
            if not target_sheet:
                logger.error("Не указан целевой лист для добавления изображения")
                return False
                
            if not os.path.exists(image_path):
                logger.error(f"Изображение не найдено: {image_path}")
                return False
            
            # Изменяем размер изображения, если нужно
            resized_img = resize_image(image_path, max_width, max_height)
            if not resized_img:
                logger.error(f"Не удалось изменить размер изображения: {image_path}")
                return False
                
            # Преобразуем в объект для Excel
            excel_img = convert_pil_to_excel_image(resized_img)
            if not excel_img:
                logger.error(f"Не удалось создать объект Excel-изображения: {image_path}")
                return False
                
            # Добавляем изображение в ячейку
            target_sheet.add_image(excel_img, cell)
            
            # Устанавливаем подходящую высоту строки и ширину столбца
            col_letter = cell[0]
            row_num = int(cell[1:])
            
            # Устанавливаем ширину столбца (1 единица ≈ 0.1640625 символа)
            col_width = excel_img.width / 7  # Приблизительное преобразование пикселей в единицы ширины столбца
            set_column_width(target_sheet, col_letter, col_width)
            
            # Устанавливаем высоту строки (1 единица ≈ 0.75 пункта)
            row_height = excel_img.height * 0.75  # Приблизительное преобразование пикселей в единицы высоты строки
            target_sheet.row_dimensions[row_num].height = row_height
            
            logger.info(f"Изображение добавлено в ячейку {cell}")
            return True
        except Exception as e:
            logger.error(f"Ошибка при добавлении изображения в ячейку {cell}: {e}")
            return False
    
    def add_images_by_article(self, image_dir: str, article_column: str, 
                             image_column: str, start_row: int = 2,
                             max_width: int = 300, max_height: int = 200,
                             sheet: Worksheet = None) -> int:
        """
        Добавляет изображения по артикулам из указанного столбца
        
        Args:
            image_dir (str): Путь к директории с изображениями
            article_column (str): Буква столбца с артикулами
            image_column (str): Буква столбца для добавления изображений
            start_row (int): Начальная строка данных
            max_width (int): Максимальная ширина изображения
            max_height (int): Максимальная высота изображения
            sheet (Worksheet, optional): Целевой лист. Если None, используется активный
            
        Returns:
            int: Количество добавленных изображений
        """
        try:
            if not self.workbook:
                logger.error("Нет активной книги Excel")
                return 0
                
            target_sheet = sheet or self.active_sheet
            if not target_sheet:
                logger.error("Не указан целевой лист для добавления изображений")
                return 0
                
            if not os.path.exists(image_dir):
                logger.error(f"Директория с изображениями не найдена: {image_dir}")
                return 0
                
            # Получаем список всех изображений в директории
            all_images = find_images_in_directory(image_dir)
            if not all_images:
                logger.warning(f"В директории {image_dir} не найдено изображений")
                return 0
                
            # Добавляем изображения для каждого артикула
            added_count = 0
            row = start_row
            
            while True:
                # Получаем значение артикула
                cell_ref = f"{article_column}{row}"
                article = get_cell_value(target_sheet, cell_ref)
                
                # Если ячейка пуста, прерываем цикл
                if not article:
                    break
                    
                # Ищем изображения для артикула
                article_str = str(article).strip().lower()
                matching_images = find_images_by_pattern(image_dir, article_str)
                
                if matching_images:
                    # Берем первое найденное изображение
                    image_path = matching_images[0]
                    
                    # Добавляем изображение в соответствующую ячейку
                    image_cell = f"{image_column}{row}"
                    if self.add_image_to_cell(image_path, image_cell, max_width, max_height, target_sheet):
                        added_count += 1
                        
                # Переходим к следующей строке
                row += 1
                
            logger.info(f"Добавлено {added_count} изображений по артикулам")
            return added_count
        except Exception as e:
            logger.error(f"Ошибка при добавлении изображений по артикулам: {e}")
            return 0
    
    def create_catalog_with_images(self, image_dir: str, output_file: str,
                                 headers: List[str] = None, 
                                 image_width: int = 300, image_height: int = 200) -> bool:
        """
        Создает каталог товаров с изображениями
        
        Args:
            image_dir (str): Путь к директории с изображениями
            output_file (str): Путь для сохранения файла каталога
            headers (List[str], optional): Заголовки столбцов
            image_width (int): Ширина изображений
            image_height (int): Высота изображений
            
        Returns:
            bool: True, если каталог успешно создан
        """
        try:
            if not os.path.exists(image_dir):
                logger.error(f"Директория с изображениями не найдена: {image_dir}")
                return False
                
            # Создаем новую рабочую книгу
            self.create_new_workbook()
            sheet = self.active_sheet
            sheet.title = "Каталог"
            
            # Устанавливаем заголовки
            if not headers:
                headers = ["Изображение", "Артикул", "Наименование", "Цена"]
            
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col).value = header
                # Применяем жирный шрифт к заголовкам
                apply_style_to_cell(sheet, f"{get_column_letter(col)}1", {
                    'font': {'bold': True},
                    'alignment': {'horizontal': 'center', 'vertical': 'center'}
                })
            
            # Получаем список изображений
            images = find_images_in_directory(image_dir)
            
            # Заполняем таблицу изображениями и информацией
            row = 2
            for img_path in images:
                # Получаем артикул из имени файла
                filename = os.path.basename(img_path)
                article = os.path.splitext(filename)[0]
                
                # Добавляем изображение
                image_cell = f"A{row}"
                self.add_image_to_cell(img_path, image_cell, image_width, image_height)
                
                # Добавляем информацию о товаре
                sheet.cell(row=row, column=2).value = article  # Артикул
                sheet.cell(row=row, column=3).value = f"Товар {article}"  # Наименование
                
                row += 1
            
            # Устанавливаем ширину столбцов
            for col, width in enumerate([40, 15, 40, 15], 1):
                set_column_width(sheet, get_column_letter(col), width)
            
            # Сохраняем файл
            return save_workbook(self.workbook, output_file)
            
        except Exception as e:
            logger.error(f"Ошибка при создании каталога с изображениями: {e}")
            return False
    
    def create_image_grid(self, images: List[str], output_file: str, 
                        cols: int = 3, image_width: int = 250, 
                        image_height: int = 200, include_names: bool = True) -> bool:
        """
        Создает сетку изображений в Excel
        
        Args:
            images (List[str]): Список путей к изображениям
            output_file (str): Путь для сохранения файла
            cols (int): Количество столбцов в сетке
            image_width (int): Ширина изображений
            image_height (int): Высота изображений
            include_names (bool): Включать ли имена файлов
            
        Returns:
            bool: True, если сетка успешно создана
        """
        try:
            if not images:
                logger.error("Не предоставлены изображения для создания сетки")
                return False
                
            # Создаем новую рабочую книгу
            self.create_new_workbook()
            sheet = self.active_sheet
            sheet.title = "Сетка изображений"
            
            # Определяем параметры сетки
            row, col = 1, 1
            row_height = image_height + 30 if include_names else image_height + 10
            
            # Устанавливаем ширину столбцов
            for c in range(1, cols+1):
                col_letter = get_column_letter(c)
                set_column_width(sheet, col_letter, image_width / 7)  # Приблизительное преобразование
            
            # Добавляем изображения
            for img_path in images:
                # Проверяем, что файл существует
                if not os.path.exists(img_path):
                    logger.warning(f"Изображение не найдено: {img_path}")
                    continue
                
                # Вычисляем позицию ячейки
                cell = f"{get_column_letter(col)}{row}"
                
                # Добавляем изображение
                self.add_image_to_cell(img_path, cell, image_width, image_height)
                
                # Добавляем имя файла, если нужно
                if include_names:
                    name_row = row + 1
                    filename = os.path.basename(img_path)
                    sheet.cell(row=name_row, column=col).value = filename
                    
                    # Центрируем текст
                    apply_style_to_cell(sheet, f"{get_column_letter(col)}{name_row}", {
                        'alignment': {'horizontal': 'center', 'vertical': 'center'}
                    })
                
                # Переходим к следующей позиции
                col += 1
                if col > cols:
                    col = 1
                    row += 2 if include_names else 1
                
                # Устанавливаем высоту строки
                sheet.row_dimensions[row].height = row_height
            
            # Сохраняем файл
            return save_workbook(self.workbook, output_file)
            
        except Exception as e:
            logger.error(f"Ошибка при создании сетки изображений: {e}")
            return False
    
    def export_images_from_excel(self, column: str, output_dir: str, 
                               start_row: int = 2, sheet: Worksheet = None) -> int:
        """
        Экспортирует изображения из Excel в указанную директорию
        
        Args:
            column (str): Буква столбца с изображениями
            output_dir (str): Директория для сохранения изображений
            start_row (int): Начальная строка
            sheet (Worksheet, optional): Лист с изображениями. Если None, используется активный
            
        Returns:
            int: Количество экспортированных изображений
        """
        try:
            if not self.workbook:
                logger.error("Нет активной книги Excel")
                return 0
                
            target_sheet = sheet or self.active_sheet
            if not target_sheet:
                logger.error("Не указан целевой лист для экспорта изображений")
                return 0
            
            # Создаем директорию для сохранения, если она не существует
            os.makedirs(output_dir, exist_ok=True)
            
            # Экспортируем изображения
            exported_count = 0
            if target_sheet._images:
                for img in target_sheet._images:
                    # Проверяем, что изображение находится в указанном столбце
                    img_col = img.anchor[0][0]  # Столбец
                    img_row = img.anchor[0][1]  # Строка
                    
                    # Проверяем, что изображение находится в нужном столбце и после начальной строки
                    if get_column_letter(img_col + 1) == column and img_row + 1 >= start_row:
                        # Определяем имя файла
                        if hasattr(img, 'filename'):
                            filename = os.path.basename(img.filename)
                        else:
                            # Если имя файла недоступно, создаем имя на основе позиции
                            filename = f"image_row{img_row + 1}.png"
                        
                        # Полный путь для сохранения
                        output_path = os.path.join(output_dir, filename)
                        
                        # Сохраняем изображение, если оно еще не существует
                        if hasattr(img, 'image') and img.image:
                            with open(output_path, "wb") as f:
                                f.write(img.image.getvalue())
                            exported_count += 1
                            logger.debug(f"Экспортировано изображение: {output_path}")
            
            logger.info(f"Экспортировано {exported_count} изображений в {output_dir}")
            return exported_count
        except Exception as e:
            logger.error(f"Ошибка при экспорте изображений: {e}")
            return 0 